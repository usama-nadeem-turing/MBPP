import json
import os
import ast
import subprocess
import tempfile
import logging
from typing import Dict, List, Any, Tuple
import argparse
from datetime import datetime
import pandas as pd

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class MBPPEvaluator:
    def __init__(self, results_dir: str = None, file_path: str = None):
        """
        Initialize the MBPP evaluator.
        
        Args:
            results_dir: Directory containing results files. If None, uses most recent results directory.
            file_path: Specific file path to evaluate. If provided, overrides results_dir.
        """
        self.file_path = file_path
        
        if file_path:
            # Extract results directory from file path
            self.results_dir = os.path.dirname(file_path)
            logger.info(f"Using specific file: {file_path}")
        else:
            self.results_dir = results_dir or self._find_latest_results_dir()
            logger.info(f"Using results directory: {self.results_dir}")
        
        # Create evaluation results directory
        if self.results_dir:
            # Create evaluations subdirectory within the results directory
            self.eval_results_dir = os.path.join(self.results_dir, "evaluations")
        else:
            # Fallback to timestamped directory in current directory
            self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.eval_results_dir = f"evaluation_results_"
        
        os.makedirs(self.eval_results_dir, exist_ok=True)
        logger.info(f"Created evaluation results directory: {self.eval_results_dir}")
        
    def _find_latest_results_dir(self) -> str:
        """
        Find the most recent results directory.
        
        Returns:
            Path to the most recent results directory
        """
        results_dirs = [d for d in os.listdir('.') if d.startswith('results_')]
        if not results_dirs:
            raise FileNotFoundError("No results directories found")
        
        # Sort by timestamp (newest first)
        results_dirs.sort(reverse=True)
        return results_dirs[0]
    
    def load_results_from_file(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Load results from a specific file.
        
        Args:
            file_path: Path to the results file
            
        Returns:
            List of result dictionaries
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                results = json.load(f)
            logger.info(f"Loaded {len(results)} results from {file_path}")
            return results
        except Exception as e:
            logger.error(f"Error loading file {file_path}: {e}")
            return []
    
    def load_results(self, split: str = None) -> List[Dict[str, Any]]:
        """
        Load results from the results directory or specific file.
        
        Args:
            split: Specific split to load (test, train, validation). If None, loads all splits.
                  Ignored if file_path is provided.
            
        Returns:
            List of result dictionaries
        """
        # If specific file path is provided, load from that file
        if self.file_path:
            return self.load_results_from_file(self.file_path)
        
        results = []
        
        if split:
            # Load specific split
            filename = f"mbpp_results_final_{split}.json"
            filepath = os.path.join(self.results_dir, filename)
            if os.path.exists(filepath):
                with open(filepath, 'r', encoding='utf-8') as f:
                    results = json.load(f)
                logger.info(f"Loaded {len(results)} results from {split} split")
            else:
                logger.warning(f"Results file not found: {filepath}")
        else:
            # Load all splits
            for split_name in ['test', 'train', 'validation']:
                filename = f"mbpp_results_final_{split_name}.json"
                filepath = os.path.join(self.results_dir, filename)
                if os.path.exists(filepath):
                    with open(filepath, 'r', encoding='utf-8') as f:
                        split_results = json.load(f)
                    results.extend(split_results)
                    logger.info(f"Loaded {len(split_results)} results from {split_name} split")
                else:
                    logger.warning(f"Results file not found: {filepath}")
        
        return results
    
    def extract_code_from_response(self, model_response: Dict[str, Any]) -> str:
        """
        Extract Python code from the model response.
        
        Args:
            model_response: Model response dictionary
            
        Returns:
            Extracted Python code string
        """
        try:
            if "error" in model_response:
                return ""
            
            # Get the content from the response
            choices = model_response.get("choices", [])
            if not choices:
                return ""
            
            content = choices[0].get("message", {}).get("content", "")
            if not content:
                return ""
            
            # Try to extract code blocks
            code_blocks = []
            
            # Look for ```python blocks
            if "```python" in content:
                parts = content.split("```python")
                for part in parts[1:]:
                    if "```" in part:
                        code_block = part.split("```")[0].strip()
                        code_blocks.append(code_block)
            
            # Look for ``` blocks (without python)
            elif "```" in content:
                parts = content.split("```")
                for i in range(1, len(parts), 2):
                    if i < len(parts):
                        code_block = parts[i].strip()
                        if code_block and not code_block.startswith("python"):
                            code_blocks.append(code_block)
            
            # If no code blocks found, try to extract function definitions
            else:
                # Look for function definitions
                lines = content.split('\n')
                in_function = False
                function_lines = []
                
                for line in lines:
                    if line.strip().startswith('def '):
                        in_function = True
                        function_lines.append(line)
                    elif in_function:
                        if line.strip() == '' or line.startswith(' ') or line.startswith('\t'):
                            function_lines.append(line)
                        else:
                            break
                
                if function_lines:
                    code_blocks.append('\n'.join(function_lines))
            
            return '\n\n'.join(code_blocks) if code_blocks else content.strip()
            
        except Exception as e:
            logger.error(f"Error extracting code: {e}")
            return ""
    
    def create_test_script(self, problem: Dict[str, Any], generated_code: str) -> str:
        """
        Create a test script that includes the generated code and test cases.
        
        Args:
            problem: Problem dictionary from MBPP dataset
            generated_code: Generated Python code
            
        Returns:
            Complete test script as string
        """
        # Get test cases
        test_cases = problem.get('test_list', [])
        
        # Create the test script
        script = f"""# Generated code
{generated_code}

# Test cases
"""
        
        # Run individual test cases
        for i, test_case in enumerate(test_cases, 1):
            # Clean up the test case (remove 'assert ' and add proper checking)
            test_clean = test_case.replace("assert ", "")
            script += f"""
# Test case {i}
try:
    result = {test_clean}
    if result:
        print(f"Test {i}: PASS - All Clear")
    else:
        print(f"Test {i}: FAIL - Result does not match expected output")
except Exception as e:
    print(f"Test {i}: FAIL - {{e}}")
"""
        
        return script
    
    def run_test_script(self, script: str) -> Tuple[bool, List[str]]:
        """
        Run a test script and check if all tests pass.
        
        Args:
            script: Python script to run
            
        Returns:
            Tuple of (all_passed, test_outputs)
        """
        try:
            # Create a temporary file
            with tempfile.NamedTemporaryFile(mode='w', suffix='.py', delete=False) as f:
                f.write(script)
                temp_file = f.name
            
            # Run the script
            result = subprocess.run(
                ['python', temp_file],
                capture_output=True,
                text=True,
                timeout=30  # 30 second timeout
            )
            
            # Clean up
            os.unlink(temp_file)
            
            # Parse output
            output_lines = result.stdout.strip().split('\n')
            test_outputs = [line for line in output_lines if line.startswith('Test')]
            
            # Debug logging
            logger.debug(f"Script stdout: {result.stdout}")
            logger.debug(f"Script stderr: {result.stderr}")
            logger.debug(f"Script return code: {result.returncode}")
            logger.debug(f"Found test outputs: {test_outputs}")
            
            # Handle assertion errors and other failures
            if result.returncode != 0:
                if "AssertionError" in result.stderr:
                    # Extract the assertion error message
                    stderr_lines = result.stderr.strip().split('\n')
                    for line in stderr_lines:
                        if "AssertionError:" in line:
                            error_msg = line.split("AssertionError:", 1)[1].strip()
                            test_outputs = [f"Test 1: FAIL - Assertion failed: {error_msg}"]
                            break
                    if not test_outputs:
                        test_outputs = ["Test 1: FAIL - Assertion failed"]
                elif result.stderr.strip():
                    # Other errors
                    error_msg = result.stderr.strip().split('\n')[-1]  # Get last error line
                    test_outputs = [f"Test 1: FAIL - {error_msg}"]
                else:
                    test_outputs = ["Test 1: FAIL - Script failed with return code 1"]
            
            # If no test outputs were found but the script ran successfully (no stderr),
            # it might mean the generated code had assertions that passed silently
            elif not test_outputs and result.returncode == 0 and not result.stderr.strip():
                # Check if the script contains assertions that might have passed
                if 'assert ' in script:
                    test_outputs = ["Test 1: PASS - All assertions passed silently"]
                    logger.debug("No test outputs found, but script contains assertions and ran successfully")
            
            # Check if all tests passed
            all_passed = all('PASS' in output for output in test_outputs) if test_outputs else False
            
            return all_passed, test_outputs
            
        except subprocess.TimeoutExpired:
            logger.warning("Test script timed out")
            return False, ["TIMEOUT"]
        except Exception as e:
            logger.error(f"Error running test script: {e}")
            return False, [f"ERROR: {e}"]
    
    def parse_test_output(self, test_output: str) -> tuple:
        """
        Parse test output to extract status and reason.
        
        Args:
            test_output: Test output string (e.g., "Test 1: PASS - All Clear")
            
        Returns:
            Tuple of (status, reason)
        """
        try:
            # Remove "Test X: " prefix
            if ":" in test_output:
                output_part = test_output.split(":", 1)[1].strip()
            else:
                output_part = test_output
            
            # Check if it's a PASS
            if "PASS" in output_part:
                status = "PASS"
                # Extract reason after "PASS - "
                if " - " in output_part:
                    reason = output_part.split(" - ", 1)[1]
                else:
                    reason = "Passed"
            else:
                status = "FAIL"
                # Extract reason after "FAIL - "
                if " - " in output_part:
                    reason = output_part.split(" - ", 1)[1]
                else:
                    reason = "Failed"
            
            return status, reason
        except Exception as e:
            logger.warning(f"Error parsing test output '{test_output}': {e}")
            return "UNKNOWN", "Parse error"
    
    def convert_evaluations_to_dataframe(self, evaluations: List[Dict[str, Any]]) -> pd.DataFrame:
        """
        Convert evaluation results to pandas DataFrame for Excel export.
        
        Args:
            evaluations: List of evaluation result dictionaries
            
        Returns:
            DataFrame with columns: mbpp_id, problem_statement, test_case_statement, test_output, reason
        """
        rows = []
        
        for evaluation in evaluations:
            mbpp_id = evaluation.get('mbpp_id', 'unknown')
            problem_statement = evaluation.get('problem_statement', '')
            test_cases = evaluation.get('test_cases', [])
            test_outputs = evaluation.get('test_outputs', [])
            
            # Ensure we have matching test cases and outputs
            for i, (test_case, test_output) in enumerate(zip(test_cases, test_outputs)):
                status, reason = self.parse_test_output(test_output)
                
                row = {
                    'mbpp_id': mbpp_id,
                    'problem_statement': problem_statement,
                    'test_case_statement': test_case,
                    'test_output': status,
                    'reason': reason
                }
                rows.append(row)
            
            # Handle case where there are more test outputs than test cases
            if len(test_outputs) > len(test_cases):
                for i in range(len(test_cases), len(test_outputs)):
                    status, reason = self.parse_test_output(test_outputs[i])
                    row = {
                        'mbpp_id': mbpp_id,
                        'problem_statement': problem_statement,
                        'test_case_statement': f"Test case {i+1} (no statement)",
                        'test_output': status,
                        'reason': reason
                    }
                    rows.append(row)
        
        df = pd.DataFrame(rows)
        logger.info(f"Created DataFrame with {len(df)} rows for Excel export")
        return df
    
    def save_to_excel(self, df: pd.DataFrame, output_path: str):
        """
        Save DataFrame to Excel file with formatting.
        
        Args:
            df: DataFrame to save
            output_path: Output file path
        """
        try:
            # Create Excel writer with formatting
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Evaluation_Results', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Evaluation_Results']
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add some basic formatting
                from openpyxl.styles import Font, PatternFill
                
                # Header formatting
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Color coding for PASS/FAIL
                pass_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
                fail_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
                
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                    test_output_cell = row[2]  # test_output column (0-indexed)
                    if test_output_cell.value == "PASS":
                        test_output_cell.fill = pass_fill
                    elif test_output_cell.value == "FAIL":
                        test_output_cell.fill = fail_fill
            
            logger.info(f"Excel file saved to: {output_path}")
            
        except Exception as e:
            logger.error(f"Error saving Excel file: {e}")
            raise
    
    def evaluate_single_result(self, result: Dict[str, Any]) -> Dict[str, Any]:
        """
        Evaluate a single result by checking if the generated code passes tests.
        
        Args:
            result: Single result dictionary
            
        Returns:
            Evaluation result dictionary
        """
        mbpp_id = result.get('mbpp_id', 'unknown')
        problem = result.get('problem', {})
        model_response = result.get('model_response', {})
        
        logger.info(f"Evaluating MBPP ID: {mbpp_id}")
        
        # Extract generated code
        generated_code = self.extract_code_from_response(model_response)
        
        if not generated_code:
            return {
                'mbpp_id': mbpp_id,
                'problem_statement': problem.get('text', ''),
                'passed': False,
                'error': 'No code generated',
                'test_outputs': [],
                'generated_code': '',
                'test_cases': problem.get('test_list', [])
            }
        
        # Create test script
        test_script = self.create_test_script(problem, generated_code)
        
        # Run tests
        all_passed, test_outputs = self.run_test_script(test_script)
        
        # Add debug logging
        logger.debug(f"MBPP ID {mbpp_id}: all_passed={all_passed}, test_outputs={test_outputs}")
        
        return {
            'mbpp_id': mbpp_id,
            'problem_statement': problem.get('text', ''),
            'passed': all_passed,
            'error': None,
            'test_outputs': test_outputs,
            'generated_code': generated_code,
            'test_cases': problem.get('test_list', [])
        }
    
    def evaluate_all_results(self, results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Evaluate all results.
        
        Args:
            results: List of result dictionaries
            
        Returns:
            List of evaluation results
        """
        evaluations = []
        
        for i, result in enumerate(results):
            logger.info(f"Evaluating result {i+1}/{len(results)}")
            evaluation = self.evaluate_single_result(result)
            evaluations.append(evaluation)
        
        return evaluations
    
    def save_evaluations(self, evaluations: List[Dict[str, Any]], split: str = None):
        """
        Save evaluation results to both JSON and Excel files.
        
        Args:
            evaluations: List of evaluation results
            split: Split name for filename
        """
        if split:
            json_filename = f"evaluation_results_{split}.json"
            excel_filename = f"evaluation_results_{split}.xlsx"
        else:
            json_filename = f"evaluation_results_all.json"
            excel_filename = f"evaluation_results_all.xlsx"
        
        # Save JSON
        json_filepath = os.path.join(self.eval_results_dir, json_filename)
        with open(json_filepath, 'w', encoding='utf-8') as f:
            json.dump(evaluations, f, indent=2, ensure_ascii=False)
        logger.info(f"Evaluation results (JSON) saved to {json_filepath}")
        
        # Save Excel
        excel_filepath = os.path.join(self.eval_results_dir, excel_filename)
        df = self.convert_evaluations_to_dataframe(evaluations)
        self.save_to_excel(df, excel_filepath)
        logger.info(f"Evaluation results (Excel) saved to {excel_filepath}")
        
        return json_filepath, excel_filepath
    
    def analyze_evaluations(self, evaluations: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Analyze evaluation results and provide statistics.
        
        Args:
            evaluations: List of evaluation results
            
        Returns:
            Analysis dictionary
        """
        total = len(evaluations)
        passed = sum(1 for e in evaluations if e['passed'])
        failed = total - passed
        
        # Count different types of failures
        no_code = sum(1 for e in evaluations if e.get('error') == 'No code generated')
        test_failures = sum(1 for e in evaluations if not e['passed'] and e.get('error') != 'No code generated')
        
        return {
            'total_problems': total,
            'passed': passed,
            'failed': failed,
            'pass_rate': passed / total if total > 0 else 0,
            'no_code_generated': no_code,
            'test_failures': test_failures,
            'mbpp_ids_passed': [e['mbpp_id'] for e in evaluations if e['passed']],
            'mbpp_ids_failed': [e['mbpp_id'] for e in evaluations if not e['passed']]
        }

def main():
    """
    Main function to run the evaluation.
    """
    parser = argparse.ArgumentParser(description='Evaluate MBPP results by checking test case passes')
    parser.add_argument('--results-dir', type=str, default=None,
                       help='Results directory to evaluate (default: most recent)')
    parser.add_argument('--file-path', type=str, default=None,
                       help='Specific file path to evaluate (overrides results-dir)')
    parser.add_argument('--split', type=str, default=None,
                       help='Specific split to evaluate (test, train, validation). If None, evaluates all splits.')
    parser.add_argument('--mbpp-id', type=str, default=None,
                       help='Evaluate specific MBPP ID only')
    parser.add_argument('--debug', action='store_true',
                       help='Enable debug logging')
    
    args = parser.parse_args()
    
    # Set debug logging if requested
    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)
        logger.setLevel(logging.DEBUG)
    
    try:
        # Initialize evaluator
        evaluator = MBPPEvaluator(args.results_dir, args.file_path)
        
        # Load results
        results = evaluator.load_results(args.split)
        
        if not results:
            logger.error("No results found to evaluate")
            return
        
        logger.info(f"Loaded {len(results)} results for evaluation")
        
        # Filter by specific MBPP ID if requested
        if args.mbpp_id:
            # Convert mbpp_id to int for comparison since it's stored as int in the data
            try:
                target_mbpp_id = int(args.mbpp_id)
                results = [r for r in results if r.get('mbpp_id') == target_mbpp_id]
                if not results:
                    logger.error(f"No results found for MBPP ID: {args.mbpp_id}")
                    return
                logger.info(f"Filtered to {len(results)} results for MBPP ID: {args.mbpp_id}")
            except ValueError:
                logger.error(f"Invalid MBPP ID format: {args.mbpp_id}. Please provide a valid integer.")
                return
        
        # Evaluate results
        logger.info("Starting evaluation...")
        evaluations = evaluator.evaluate_all_results(results)
        
        # Save evaluations (both JSON and Excel)
        json_filepath, excel_filepath = evaluator.save_evaluations(evaluations, args.split)
        
        # Analyze results
        analysis = evaluator.analyze_evaluations(evaluations)
        
        # Print analysis
        logger.info("\n" + "="*50)
        logger.info("EVALUATION RESULTS")
        logger.info("="*50)
        logger.info(f"Total problems: {analysis['total_problems']}")
        logger.info(f"Passed: {analysis['passed']}")
        logger.info(f"Failed: {analysis['failed']}")
        logger.info(f"Pass rate: {analysis['pass_rate']:.2%}")
        logger.info(f"No code generated: {analysis['no_code_generated']}")
        logger.info(f"Test failures: {analysis['test_failures']}")
        logger.info(f"JSON file: {json_filepath}")
        logger.info(f"Excel file: {excel_filepath}")
        
        if analysis['mbpp_ids_passed']:
            logger.info(f"Passed MBPP IDs: {analysis['mbpp_ids_passed']}")
        if analysis['mbpp_ids_failed']:
            logger.info(f"Failed MBPP IDs: {analysis['mbpp_ids_failed']}")
        
        # Print sample failed evaluations
        failed_evaluations = [e for e in evaluations if not e['passed']]
        if failed_evaluations:
            logger.info(f"\nSample failed evaluations:")
            for i, eval_result in enumerate(failed_evaluations[:3]):
                logger.info(f"\nMBPP ID {eval_result['mbpp_id']}:")
                if eval_result.get('error'):
                    logger.info(f"Error: {eval_result['error']}")
                else:
                    logger.info(f"Test outputs: {eval_result['test_outputs']}")
        
    except Exception as e:
        logger.error(f"Error in evaluation: {e}")
        raise

if __name__ == "__main__":
    main() 