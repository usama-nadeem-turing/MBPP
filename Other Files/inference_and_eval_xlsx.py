import requests
import json
import time
import os
import argparse
import ast
import subprocess
import tempfile
from typing import List, Dict, Any, Tuple
from datasets import load_dataset
import logging
from datetime import datetime
import pandas as pd

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class MBPPInference:
    def __init__(self, model_url: str = "http://localhost:18000/v1/chat/completions", 
                 model_name: str = "Qwen/Qwen2.5-1.5B-Instruct"):
        """
        Initialize the MBPP inference class.
        
        Args:
            model_url: URL of the local model server
            model_name: Name of the model to use
        """
        self.model_url = model_url
        self.model_name = model_name
        self.session = requests.Session()
        
        # Create results directory with timestamp
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.results_dir = f"results_{self.timestamp}"
        os.makedirs(self.results_dir, exist_ok=True)
        logger.info(f"Created results directory: {self.results_dir}")
        
        # Create evaluation results directory
        self.eval_results_dir = f"evaluation_results_{self.timestamp}"
        os.makedirs(self.eval_results_dir, exist_ok=True)
        logger.info(f"Created evaluation results directory: {self.eval_results_dir}")
        
    def _find_latest_results_dir(self) -> str:
        """
        Find the most recent results directory.
        
        Returns:
            Path to the most recent results directory
        """
        results_dirs = [d for d in os.listdir('.') if d.startswith('results_')]
        if not results_dirs:
            raise FileNotFoundError("No results directories found")
        
        # Sort by timestamp (newest first)
        results_dirs.sort(reverse=True)
        return results_dirs[0]
    
    def load_mbpp_dataset(self, split: str = "test") -> List[Dict[str, Any]]:
        """
        Load the MBPP dataset from Hugging Face.
        
        Args:
            split: Dataset split to load ('train', 'validation', 'test')
            
        Returns:
            List of problem dictionaries
        """
        try:
            logger.info(f"Loading MBPP dataset with split: {split}")
            dataset = load_dataset("mbpp")
            
            # Log available splits
            available_splits = list(dataset.keys())
            logger.info(f"Available splits in MBPP dataset: {available_splits}")
            
            if split not in dataset:
                logger.warning(f"Split '{split}' not found. Available splits: {available_splits}")
                split = available_splits[0]  # Use first available split
                logger.info(f"Using fallback split: {split}")
            else:
                logger.info(f"Successfully using requested split: {split}")
                
            problems = dataset[split]
            logger.info(f"Loaded {len(problems)} problems from MBPP {split} split")
            
            # Debug: Check the type of problems
            logger.debug(f"Problems type: {type(problems)}")
            logger.debug(f"Problems is list: {isinstance(problems, list)}")
            
            # Convert to list if it's not already
            if not isinstance(problems, list):
                logger.warning(f"Problems is not a list, converting from {type(problems)}")
                problems = list(problems)
            
            # Inspect the structure of the first problem
            if len(problems) > 0:
                first_problem = problems[0]
                logger.info(f"First problem structure - Keys: {list(first_problem.keys())}")
                logger.info(f"Sample problem: {first_problem}")
                logger.debug(f"First problem type: {type(first_problem)}")
            
            return problems
            
        except Exception as e:
            logger.error(f"Error loading MBPP dataset: {e}")
            raise
    
    def inspect_dataset_structure(self, problems: List[Dict[str, Any]], num_samples: int = 3):
        """
        Inspect the structure of the dataset to understand available fields.
        
        Args:
            problems: List of problem dictionaries
            num_samples: Number of sample problems to inspect
        """
        logger.info("=== DATASET STRUCTURE INSPECTION ===")
        for i in range(min(num_samples, len(problems))):
            problem = problems[i]
            logger.info(f"\nProblem {i+1} (Task ID: {problem.get('task_id', 'N/A')}):")
            logger.info(f"Keys: {list(problem.keys())}")
            for key, value in problem.items():
                if isinstance(value, str) and len(value) > 100:
                    logger.info(f"{key}: {value[:100]}...")
                else:
                    logger.info(f"{key}: {value}")
        logger.info("=== END INSPECTION ===\n")
    
    def create_prompt(self, problem: Dict[str, Any]) -> str:
        """
        Create a prompt for the model based on the problem.
        
        Args:
            problem: Problem dictionary from MBPP dataset
            
        Returns:
            Formatted prompt string
        """
        # Debug: Print available keys to understand the structure
        logger.debug(f"Available keys in problem: {list(problem.keys())}")
        
        # Get test cases if available
        test_cases = ""
        if 'test_list' in problem and problem['test_list']:
            test_cases = "\n\nExpected behavior (test cases):\n"
            for i, test in enumerate(problem['test_list'], 1):
                # Clean up the assert statement to show the expected behavior
                test_clean = test.replace("assert ", "").replace(" == ", " should return ")
                test_cases += f"{i}. {test_clean}\n"
        
        prompt = f"""Please solve the following Python programming problem:

Problem: {problem['text']}

Task ID: {problem['task_id']}{test_cases}

Please provide a complete Python function that solves this problem. Write only the function code without any explanations or comments."""
        
        return prompt
    
    def get_model_response(self, prompt: str, temperature: float = 1.0, 
                          max_tokens: int = 512) -> Dict[str, Any]:
        """
        Get response from the local model server.
        
        Args:
            prompt: Input prompt
            temperature: Sampling temperature
            max_tokens: Maximum tokens to generate
            
        Returns:
            Model response dictionary
        """
        payload = {
            "model": self.model_name,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": temperature,
            "max_tokens": max_tokens
        }
        
        try:
            response = self.session.post(
                self.model_url,
                headers={"Content-Type": "application/json"},
                json=payload,
                timeout=60
            )
            response.raise_for_status()
            return response.json()
            
        except requests.exceptions.RequestException as e:
            logger.error(f"Error calling model API: {e}")
            return {"error": str(e)}
    
    def process_problems(self, problems: List[Dict[str, Any]], 
                        max_problems: int = None, 
                        save_results: bool = True,
                        demo_mode: bool = False,
                        split_name: str = "",
                        temperature: float = 1.0,
                        max_tokens: int = 512) -> List[Dict[str, Any]]:
        """
        Process all problems and get model responses.
        
        Args:
            problems: List of problem dictionaries
            max_problems: Maximum number of problems to process (None for all)
            save_results: Whether to save results to file
            demo_mode: If True, process only 3-4 problems for demo
            split_name: Name of the split being processed
            temperature: Sampling temperature for model generation
            max_tokens: Maximum tokens to generate
            
        Returns:
            List of results with problems and model responses
        """
        if demo_mode:
            max_problems = 4
            logger.info(f"Running in DEMO mode - processing 4 problems only for split: {split_name}")
        
        # Debug: Check the type and structure of problems
        logger.debug(f"Problems type: {type(problems)}")
        logger.debug(f"Problems length: {len(problems)}")
        if len(problems) > 0:
            logger.debug(f"First problem type: {type(problems[0])}")
            logger.debug(f"First problem: {problems[0]}")
        
        # Apply max_problems limit by slicing the problems list
        if max_problems:
            problems = problems[:max_problems]
            logger.info(f"Limited to {len(problems)} problems for split: {split_name}")
            logger.debug(f"After slicing - First problem type: {type(problems[0]) if problems else 'No problems'}")
            
        results = []
        
        for i, problem in enumerate(problems):
            # Debug: Check if problem is a dictionary
            if not isinstance(problem, dict):
                logger.error(f"Problem {i} is not a dictionary: {type(problem)} - {problem}")
                continue
                
            logger.info(f"Processing problem {i+1}/{len(problems)}: Task {problem.get('task_id', 'unknown')} for split: {split_name}")
            
            # Create prompt
            prompt = self.create_prompt(problem)
            
            # Get model response with temperature and max_tokens
            model_response = self.get_model_response(prompt, temperature=temperature, max_tokens=max_tokens)
            
            # Store result with MBPP ID
            result = {
                "mbpp_id": problem.get('task_id', f'unknown_{i}'),
                "problem": problem,
                "prompt": prompt,
                "model_response": model_response,
                "timestamp": time.time()
            }
            
            results.append(result)
            
            # Save result incrementally after each task
            if save_results:
                if demo_mode:
                    filename = f"mbpp_demo_results_{split_name}.json"
                else:
                    filename = f"mbpp_results_final_{split_name}.json"
                self.save_results(results, filename)
                logger.info(f"Saved incremental results after task {i+1} to {filename}")
            
            # Add delay to avoid overwhelming the server
            time.sleep(0.5)
            
        logger.info(f"Completed processing {len(results)} problems for split: {split_name}")
        return results
    
    def save_results(self, results: List[Dict[str, Any]], filename: str):
        """
        Save results to a JSON file in the results directory.
        
        Args:
            results: List of result dictionaries
            filename: Output filename
        """
        try:
            filepath = os.path.join(self.results_dir, filename)
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2, ensure_ascii=False)
            logger.info(f"Results saved to {filepath}")
        except Exception as e:
            logger.error(f"Error saving results: {e}")
    
    def save_run_metadata(self, args, problems_count: int, demo_mode: bool = False, split: str = ""):
        """
        Save run metadata and configuration to the results directory.
        
        Args:
            args: Command line arguments
            problems_count: Number of problems to be processed
            demo_mode: Whether running in demo mode
            split: Name of the split being processed
        """
        try:
            metadata = {
                "timestamp": self.timestamp,
                "run_config": {
                    "demo_mode": demo_mode,
                    "split": split,
                    "max_problems": args.max_problems,
                    "model_url": args.model_url,
                    "model_name": args.model_name,
                    "temperature": args.temperature,
                    "max_tokens": args.max_tokens,
                    "debug": args.debug
                },
                "dataset_info": {
                    "total_problems_loaded": problems_count,
                    "problems_to_process": min(problems_count, args.max_problems or problems_count) if not demo_mode else 4
                },
                "results_directory": self.results_dir
            }
            
            filepath = os.path.join(self.results_dir, f"run_metadata_{split}.json")
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(metadata, f, indent=2, ensure_ascii=False)
            logger.info(f"Run metadata saved to {filepath}")
        except Exception as e:
            logger.error(f"Error saving run metadata: {e}")
    
    def save_analysis(self, analysis: Dict[str, Any], demo_mode: bool = False, split: str = ""):
        """
        Save analysis results to the results directory.
        
        Args:
            analysis: Analysis dictionary
            demo_mode: Whether running in demo mode
            split: Name of the split being processed
        """
        try:
            if demo_mode:
                filename = f"analysis_demo_{split}.json"
            else:
                filename = f"analysis_final_{split}.json"
            
            filepath = os.path.join(self.results_dir, filename)
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(analysis, f, indent=2, ensure_ascii=False)
            logger.info(f"Analysis saved to {filepath}")
        except Exception as e:
            logger.error(f"Error saving analysis: {e}")
    
    def analyze_results(self, results: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Analyze the results and provide statistics.
        
        Args:
            results: List of result dictionaries
            
        Returns:
            Analysis dictionary with statistics
        """
        total_problems = len(results)
        successful_responses = sum(1 for r in results if "error" not in r["model_response"])
        failed_responses = total_problems - successful_responses
        
        # Calculate average response time (if available)
        response_times = []
        for r in results:
            if "model_response" in r and "usage" in r["model_response"]:
                usage = r["model_response"]["usage"]
                if "total_tokens" in usage:
                    response_times.append(usage["total_tokens"])
        
        analysis = {
            "total_problems": total_problems,
            "successful_responses": successful_responses,
            "failed_responses": failed_responses,
            "success_rate": successful_responses / total_problems if total_problems > 0 else 0,
            "average_tokens": sum(response_times) / len(response_times) if response_times else 0,
            "total_tokens": sum(response_times) if response_times else 0,
            "mbpp_ids": [r["mbpp_id"] for r in results]
        }
        
        return analysis

    # ==================== EVALUATION METHODS ====================
    
    def extract_code_from_response(self, model_response: Dict[str, Any]) -> str:
        """
        Extract Python code from the model response.
        
        Args:
            model_response: Model response dictionary
            
        Returns:
            Extracted Python code string
        """
        try:
            if "error" in model_response:
                return ""
            
            # Get the content from the response
            choices = model_response.get("choices", [])
            if not choices:
                return ""
            
            content = choices[0].get("message", {}).get("content", "")
            if not content:
                return ""
            
            # Try to extract code blocks
            code_blocks = []
            
            # Look for ```python blocks
            if "```python" in content:
                parts = content.split("```python")
                for part in parts[1:]:
                    if "```" in part:
                        code_block = part.split("```")[0].strip()
                        code_blocks.append(code_block)
            
            # Look for ``` blocks (without python)
            elif "```" in content:
                parts = content.split("```")
                for i in range(1, len(parts), 2):
                    if i < len(parts):
                        code_block = parts[i].strip()
                        if code_block and not code_block.startswith("python"):
                            code_blocks.append(code_block)
            
            # If no code blocks found, try to extract function definitions
            else:
                # Look for function definitions
                lines = content.split('\n')
                in_function = False
                function_lines = []
                
                for line in lines:
                    if line.strip().startswith('def '):
                        in_function = True
                        function_lines.append(line)
                    elif in_function:
                        if line.strip() == '' or line.startswith(' ') or line.startswith('\t'):
                            function_lines.append(line)
                        else:
                            break
                
                if function_lines:
                    code_blocks.append('\n'.join(function_lines))
            
            return '\n\n'.join(code_blocks) if code_blocks else content.strip()
            
        except Exception as e:
            logger.error(f"Error extracting code: {e}")
            return ""
    
    def create_test_script(self, problem: Dict[str, Any], generated_code: str) -> str:
        """
        Create a test script that includes the generated code and test cases.
        
        Args:
            problem: Problem dictionary from MBPP dataset
            generated_code: Generated Python code
            
        Returns:
            Complete test script as string
        """
        # Get test cases
        test_cases = problem.get('test_list', [])
        
        # Create the test script
        script = f"""# Generated code
{generated_code}

# Test cases
"""
        
        for i, test_case in enumerate(test_cases, 1):
            # Clean up the test case (remove 'assert ' and add proper checking)
            test_clean = test_case.replace("assert ", "")
            script += f"""
# Test case {i}
try:
    result = {test_clean}
    if result:
        print(f"Test {i}: PASS - All Clear")
    else:
        print(f"Test {i}: FAIL - Result does not match expected output")
except Exception as e:
    print(f"Test {i}: FAIL - {{e}}")
"""
        
        return script
    
    def run_test_script(self, script: str) -> Tuple[bool, List[str]]:
        """
        Run a test script and check if all tests pass.
        
        Args:
            script: Python script to run
            
        Returns:
            Tuple of (all_passed, test_outputs)
        """
        try:
            # Create a temporary file
            with tempfile.NamedTemporaryFile(mode='w', suffix='.py', delete=False) as f:
                f.write(script)
                temp_file = f.name
            
            # Run the script
            result = subprocess.run(
                ['python', temp_file],
                capture_output=True,
                text=True,
                timeout=30  # 30 second timeout
            )
            
            # Clean up
            os.unlink(temp_file)
            
            # Parse output
            output_lines = result.stdout.strip().split('\n')
            test_outputs = [line for line in output_lines if line.startswith('Test')]
            
            # Check if all tests passed
            all_passed = all('PASS' in output for output in test_outputs)
            
            return all_passed, test_outputs
            
        except subprocess.TimeoutExpired:
            logger.warning("Test script timed out")
            return False, ["TIMEOUT"]
        except Exception as e:
            logger.error(f"Error running test script: {e}")
            return False, [f"ERROR: {e}"]
    
    def parse_test_output(self, test_output: str) -> tuple:
        """
        Parse test output to extract status and reason.
        
        Args:
            test_output: Test output string (e.g., "Test 1: PASS - All Clear")
            
        Returns:
            Tuple of (status, reason)
        """
        try:
            # Remove "Test X: " prefix
            if ":" in test_output:
                output_part = test_output.split(":", 1)[1].strip()
            else:
                output_part = test_output
            
            # Check if it's a PASS
            if "PASS" in output_part:
                status = "PASS"
                # Extract reason after "PASS - "
                if " - " in output_part:
                    reason = output_part.split(" - ", 1)[1]
                else:
                    reason = "Passed"
            else:
                status = "FAIL"
                # Extract reason after "FAIL - "
                if " - " in output_part:
                    reason = output_part.split(" - ", 1)[1]
                else:
                    reason = "Failed"
            
            return status, reason
        except Exception as e:
            logger.warning(f"Error parsing test output '{test_output}': {e}")
            return "UNKNOWN", "Parse error"
    
    def convert_evaluations_to_dataframe(self, evaluations: List[Dict[str, Any]]) -> pd.DataFrame:
        """
        Convert evaluation results to pandas DataFrame for Excel export.
        
        Args:
            evaluations: List of evaluation result dictionaries
            
        Returns:
            DataFrame with columns: mbpp_id, test_case_statement, test_output, reason
        """
        rows = []
        
        for evaluation in evaluations:
            mbpp_id = evaluation.get('mbpp_id', 'unknown')
            test_cases = evaluation.get('test_cases', [])
            test_outputs = evaluation.get('test_outputs', [])
            
            # Ensure we have matching test cases and outputs
            for i, (test_case, test_output) in enumerate(zip(test_cases, test_outputs)):
                status, reason = self.parse_test_output(test_output)
                
                row = {
                    'mbpp_id': mbpp_id,
                    'test_case_statement': test_case,
                    'test_output': status,
                    'reason': reason
                }
                rows.append(row)
            
            # Handle case where there are more test outputs than test cases
            if len(test_outputs) > len(test_cases):
                for i in range(len(test_cases), len(test_outputs)):
                    status, reason = self.parse_test_output(test_outputs[i])
                    row = {
                        'mbpp_id': mbpp_id,
                        'test_case_statement': f"Test case {i+1} (no statement)",
                        'test_output': status,
                        'reason': reason
                    }
                    rows.append(row)
        
        df = pd.DataFrame(rows)
        logger.info(f"Created DataFrame with {len(df)} rows for Excel export")
        return df
    
    def save_to_excel(self, df: pd.DataFrame, output_path: str):
        """
        Save DataFrame to Excel file with formatting.
        
        Args:
            df: DataFrame to save
            output_path: Output file path
        """
        try:
            # Create Excel writer with formatting
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Evaluation_Results', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Evaluation_Results']
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add some basic formatting
                from openpyxl.styles import Font, PatternFill
                
                # Header formatting
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Color coding for PASS/FAIL
                pass_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
                fail_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
                
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                    test_output_cell = row[2]  # test_output column (0-indexed)
                    if test_output_cell.value == "PASS":
                        test_output_cell.fill = pass_fill
                    elif test_output_cell.value == "FAIL":
                        test_output_cell.fill = fail_fill
            
            logger.info(f"Excel file saved to: {output_path}")
            
        except Exception as e:
            logger.error(f"Error saving Excel file: {e}")
            raise
    
    def evaluate_single_result(self, result: Dict[str, Any]) -> Dict[str, Any]:
        """
        Evaluate a single result by checking if the generated code passes tests.
        
        Args:
            result: Single result dictionary
            
        Returns:
            Evaluation result dictionary
        """
        mbpp_id = result.get('mbpp_id', 'unknown')
        problem = result.get('problem', {})
        model_response = result.get('model_response', {})
        
        logger.info(f"Evaluating MBPP ID: {mbpp_id}")
        
        # Extract generated code
        generated_code = self.extract_code_from_response(model_response)
        
        if not generated_code:
            return {
                'mbpp_id': mbpp_id,
                'passed': False,
                'error': 'No code generated',
                'test_outputs': [],
                'generated_code': '',
                'test_cases': problem.get('test_list', [])
            }
        
        # Create test script
        test_script = self.create_test_script(problem, generated_code)
        
        # Run tests
        all_passed, test_outputs = self.run_test_script(test_script)
        
        return {
            'mbpp_id': mbpp_id,
            'passed': all_passed,
            'error': None,
            'test_outputs': test_outputs,
            'generated_code': generated_code,
            'test_cases': problem.get('test_list', [])
        }
    
    def evaluate_all_results(self, results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Evaluate all results.
        
        Args:
            results: List of result dictionaries
            
        Returns:
            List of evaluation results
        """
        evaluations = []
        
        for i, result in enumerate(results):
            logger.info(f"Evaluating result {i+1}/{len(results)}")
            evaluation = self.evaluate_single_result(result)
            evaluations.append(evaluation)
        
        return evaluations
    
    def save_evaluations(self, evaluations: List[Dict[str, Any]], split: str = None):
        """
        Save evaluation results to both JSON and Excel files.
        
        Args:
            evaluations: List of evaluation results
            split: Split name for filename
        """
        if split:
            json_filename = f"evaluation_results_{split}.json"
            excel_filename = f"evaluation_results_{split}.xlsx"
        else:
            json_filename = f"evaluation_results_all.json"
            excel_filename = f"evaluation_results_all.xlsx"
        
        # Save JSON
        json_filepath = os.path.join(self.eval_results_dir, json_filename)
        with open(json_filepath, 'w', encoding='utf-8') as f:
            json.dump(evaluations, f, indent=2, ensure_ascii=False)
        logger.info(f"Evaluation results (JSON) saved to {json_filepath}")
        
        # Save Excel
        excel_filepath = os.path.join(self.eval_results_dir, excel_filename)
        df = self.convert_evaluations_to_dataframe(evaluations)
        self.save_to_excel(df, excel_filepath)
        logger.info(f"Evaluation results (Excel) saved to {excel_filepath}")
        
        return json_filepath, excel_filepath
    
    def analyze_evaluations(self, evaluations: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Analyze evaluation results and provide statistics.
        
        Args:
            evaluations: List of evaluation results
            
        Returns:
            Analysis dictionary
        """
        total = len(evaluations)
        passed = sum(1 for e in evaluations if e['passed'])
        failed = total - passed
        
        # Count different types of failures
        no_code = sum(1 for e in evaluations if e.get('error') == 'No code generated')
        test_failures = sum(1 for e in evaluations if not e['passed'] and e.get('error') != 'No code generated')
        
        return {
            'total_problems': total,
            'passed': passed,
            'failed': failed,
            'pass_rate': passed / total if total > 0 else 0,
            'no_code_generated': no_code,
            'test_failures': test_failures,
            'mbpp_ids_passed': [e['mbpp_id'] for e in evaluations if e['passed']],
            'mbpp_ids_failed': [e['mbpp_id'] for e in evaluations if not e['passed']]
        }

def main():
    """
    Main function to run the MBPP inference and evaluation.
    """
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='Run MBPP inference with local model server and evaluation')
    parser.add_argument('--demo', action='store_true', 
                       help='Run in demo mode with only 3-4 problems')
    parser.add_argument('--split', type=str, default='test',
                       help='Dataset split to use (train, validation, test). Use "all" to process all splits.')
    parser.add_argument('--max-problems', type=int, default=None,
                       help='Maximum number of problems to process')
    parser.add_argument('--model-url', type=str, 
                       default='http://localhost:18000/v1/chat/completions',
                       help='URL of the model server')
    parser.add_argument('--model-name', type=str, 
                       default='Qwen/Qwen2.5-1.5B-Instruct',
                       help='Name of the model to use')
    parser.add_argument('--temperature', type=float, default=1.0,
                       help='Sampling temperature')
    parser.add_argument('--max-tokens', type=int, default=512,
                       help='Maximum tokens to generate')
    parser.add_argument('--debug', action='store_true',
                       help='Enable debug logging')
    # Evaluation arguments
    parser.add_argument('--evaluate', action='store_true',
                       help='Run evaluation after inference')
    parser.add_argument('--evaluate-only', action='store_true',
                       help='Only run evaluation on existing results (skip inference)')
    parser.add_argument('--results-dir', type=str, default=None,
                       help='Results directory to evaluate (default: most recent)')
    parser.add_argument('--file-path', type=str, default=None,
                       help='Specific file path to evaluate (overrides results-dir)')
    parser.add_argument('--mbpp-id', type=str, default=None,
                       help='Evaluate specific MBPP ID only')
    
    args = parser.parse_args()
    
    # Set debug logging if requested
    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)
        logger.setLevel(logging.DEBUG)
    
    # Initialize the inference class
    inference = MBPPInference(
        model_url=args.model_url,
        model_name=args.model_name
    )
    
    try:
        # If evaluate-only mode, skip inference and go directly to evaluation
        if args.evaluate_only:
            logger.info("Running in EVALUATE-ONLY mode")
            
            # Load results for evaluation
            if args.file_path:
                # Load from specific file
                with open(args.file_path, 'r', encoding='utf-8') as f:
                    results = json.load(f)
                logger.info(f"Loaded {len(results)} results from {args.file_path}")
            else:
                # Load from results directory
                results_dir = args.results_dir or inference._find_latest_results_dir()
                logger.info(f"Using results directory: {results_dir}")
                
                # Load all splits or specific split
                if args.split.lower() == 'all':
                    results = []
                    for split_name in ['test', 'train', 'validation']:
                        filename = f"mbpp_results_final_{split_name}.json"
                        filepath = os.path.join(results_dir, filename)
                        if os.path.exists(filepath):
                            with open(filepath, 'r', encoding='utf-8') as f:
                                split_results = json.load(f)
                            results.extend(split_results)
                            logger.info(f"Loaded {len(split_results)} results from {split_name} split")
                else:
                    filename = f"mbpp_results_final_{args.split}.json"
                    filepath = os.path.join(results_dir, filename)
                    if os.path.exists(filepath):
                        with open(filepath, 'r', encoding='utf-8') as f:
                            results = json.load(f)
                        logger.info(f"Loaded {len(results)} results from {args.split} split")
                    else:
                        logger.error(f"Results file not found: {filepath}")
                        return
            
            # Filter by specific MBPP ID if requested
            if args.mbpp_id:
                results = [r for r in results if r.get('mbpp_id') == args.mbpp_id]
                if not results:
                    logger.error(f"No results found for MBPP ID: {args.mbpp_id}")
                    return
                logger.info(f"Filtered to {len(results)} results for MBPP ID: {args.mbpp_id}")
            
            # Run evaluation
            logger.info("Starting evaluation...")
            evaluations = inference.evaluate_all_results(results)
            
            # Save evaluations
            json_filepath, excel_filepath = inference.save_evaluations(evaluations, args.split)
            
            # Analyze evaluation results
            eval_analysis = inference.analyze_evaluations(evaluations)
            
            # Print evaluation analysis
            logger.info("\n" + "="*50)
            logger.info("EVALUATION RESULTS")
            logger.info("="*50)
            logger.info(f"Total problems: {eval_analysis['total_problems']}")
            logger.info(f"Passed: {eval_analysis['passed']}")
            logger.info(f"Failed: {eval_analysis['failed']}")
            logger.info(f"Pass rate: {eval_analysis['pass_rate']:.2%}")
            logger.info(f"No code generated: {eval_analysis['no_code_generated']}")
            logger.info(f"Test failures: {eval_analysis['test_failures']}")
            logger.info(f"JSON file: {json_filepath}")
            logger.info(f"Excel file: {excel_filepath}")
            
            return
        
        # Load MBPP dataset to get available splits
        logger.info("Loading MBPP dataset to check available splits...")
        temp_dataset = load_dataset("mbpp")
        available_splits = list(temp_dataset.keys())
        logger.info(f"Available splits in MBPP dataset: {available_splits}")
        
        # Determine which splits to process
        if args.split.lower() == 'all':
            # Exclude 'prompt' split as it's not typically used for evaluation
            splits_to_process = [split for split in available_splits if split != 'prompt']
            logger.info(f"Processing all splits (excluding 'prompt'): {splits_to_process}")
        else:
            if args.split not in available_splits:
                logger.error(f"Split '{args.split}' not found. Available splits: {available_splits}")
                return
            splits_to_process = [args.split]
            logger.info(f"Processing specific split: {args.split}")
        
        all_results = {}
        all_analyses = {}
        all_evaluations = {}
        all_eval_analyses = {}
        
        # Process each split
        for split in splits_to_process:
            logger.info(f"\n{'='*50}")
            logger.info(f"Processing split: {split}")
            logger.info(f"{'='*50}")
            
            # Load problems for this split
            problems = inference.load_mbpp_dataset(split=split)
            
            # Inspect dataset structure
            inference.inspect_dataset_structure(problems)
            
            # Save run metadata for this split
            inference.save_run_metadata(args, len(problems), args.demo, split)
            
            # Process problems
            if args.demo:
                logger.info(f"Starting DEMO mode for split {split} - processing 4 problems...")
                results = inference.process_problems(
                    problems, 
                    save_results=True,
                    demo_mode=True,
                    split_name=split,
                    temperature=args.temperature,
                    max_tokens=args.max_tokens
                )
            else:
                logger.info(f"Starting to process problems for split {split}...")
                results = inference.process_problems(
                    problems, 
                    max_problems=args.max_problems, 
                    save_results=True,
                    demo_mode=False,
                    split_name=split,
                    temperature=args.temperature,
                    max_tokens=args.max_tokens
                )
            
            # Analyze results for this split
            analysis = inference.analyze_results(results)
            all_results[split] = results
            all_analyses[split] = analysis
            
            # Save analysis for this split
            inference.save_analysis(analysis, args.demo, split)
            
            # Print analysis for this split
            logger.info(f"\n=== ANALYSIS RESULTS FOR SPLIT: {split} ===")
            for key, value in analysis.items():
                logger.info(f"{key}: {value}")
            
            # Run evaluation if requested
            if args.evaluate:
                logger.info(f"\n=== EVALUATING RESULTS FOR SPLIT: {split} ===")
                evaluations = inference.evaluate_all_results(results)
                all_evaluations[split] = evaluations
                
                # Save evaluations
                json_filepath, excel_filepath = inference.save_evaluations(evaluations, split)
                
                # Analyze evaluation results
                eval_analysis = inference.analyze_evaluations(evaluations)
                all_eval_analyses[split] = eval_analysis
                
                # Print evaluation analysis
                logger.info(f"\n=== EVALUATION RESULTS FOR SPLIT: {split} ===")
                logger.info(f"Total problems: {eval_analysis['total_problems']}")
                logger.info(f"Passed: {eval_analysis['passed']}")
                logger.info(f"Failed: {eval_analysis['failed']}")
                logger.info(f"Pass rate: {eval_analysis['pass_rate']:.2%}")
                logger.info(f"No code generated: {eval_analysis['no_code_generated']}")
                logger.info(f"Test failures: {eval_analysis['test_failures']}")
                logger.info(f"JSON file: {json_filepath}")
                logger.info(f"Excel file: {excel_filepath}")
            
            # Print sample results for this split
            logger.info(f"\n=== SAMPLE RESULTS FOR SPLIT: {split} ===")
            for i, result in enumerate(results[:3]):
                logger.info(f"\nProblem {i+1} (MBPP ID: {result['mbpp_id']}):")
                logger.info(f"Prompt: {result['prompt'][:200]}...")
                
                if "error" in result["model_response"]:
                    logger.info(f"Error: {result['model_response']['error']}")
                else:
                    response_content = result["model_response"].get("choices", [{}])[0].get("message", {}).get("content", "")
                    logger.info(f"Response: {response_content[:200]}...")
        
        # Print overall summary if processing multiple splits
        if len(splits_to_process) > 1:
            logger.info(f"\n{'='*50}")
            logger.info("OVERALL SUMMARY")
            logger.info(f"{'='*50}")
            total_problems = sum(len(results) for results in all_results.values())
            total_successful = sum(analysis['successful_responses'] for analysis in all_analyses.values())
            overall_success_rate = total_successful / total_problems if total_problems > 0 else 0
            
            logger.info(f"Total problems processed: {total_problems}")
            logger.info(f"Total successful responses: {total_successful}")
            logger.info(f"Overall success rate: {overall_success_rate:.2%}")
            
            for split in splits_to_process:
                analysis = all_analyses[split]
                logger.info(f"{split}: {analysis['successful_responses']}/{analysis['total_problems']} successful ({analysis['success_rate']:.2%})")
            
            # Print overall evaluation summary if evaluation was run
            if args.evaluate and all_eval_analyses:
                logger.info(f"\n{'='*50}")
                logger.info("OVERALL EVALUATION SUMMARY")
                logger.info(f"{'='*50}")
                total_eval_problems = sum(analysis['total_problems'] for analysis in all_eval_analyses.values())
                total_passed = sum(analysis['passed'] for analysis in all_eval_analyses.values())
                overall_pass_rate = total_passed / total_eval_problems if total_eval_problems > 0 else 0
                
                logger.info(f"Total problems evaluated: {total_eval_problems}")
                logger.info(f"Total passed: {total_passed}")
                logger.info(f"Overall pass rate: {overall_pass_rate:.2%}")
                
                for split in splits_to_process:
                    if split in all_eval_analyses:
                        eval_analysis = all_eval_analyses[split]
                        logger.info(f"{split}: {eval_analysis['passed']}/{eval_analysis['total_problems']} passed ({eval_analysis['pass_rate']:.2%})")
        
    except Exception as e:
        logger.error(f"Error in main execution: {e}")
        raise

if __name__ == "__main__":
    main()
