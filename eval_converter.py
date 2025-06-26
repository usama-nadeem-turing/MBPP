import json
import pandas as pd
import argparse
import os
import logging
from typing import List, Dict, Any
from datetime import datetime

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class EvalConverter:
    def __init__(self, file_path: str):
        """
        Initialize the evaluation converter.
        
        Args:
            file_path: Path to the evaluation results JSON file
        """
        self.file_path = file_path
        logger.info(f"Initializing converter with file: {file_path}")
    
    def load_evaluation_results(self) -> List[Dict[str, Any]]:
        """
        Load evaluation results from JSON file.
        
        Returns:
            List of evaluation result dictionaries
        """
        try:
            with open(self.file_path, 'r', encoding='utf-8') as f:
                results = json.load(f)
            logger.info(f"Loaded {len(results)} evaluation results from {self.file_path}")
            return results
        except Exception as e:
            logger.error(f"Error loading evaluation results: {e}")
            raise
    
    def parse_test_output(self, test_output: str) -> tuple:
        """
        Parse test output to extract status and reason.
        
        Args:
            test_output: Test output string (e.g., "Test 1: PASS - All Clear")
            
        Returns:
            Tuple of (status, reason)
        """
        try:
            # Remove "Test X: " prefix
            if ":" in test_output:
                output_part = test_output.split(":", 1)[1].strip()
            else:
                output_part = test_output
            
            # Check if it's a PASS
            if "PASS" in output_part:
                status = "PASS"
                # Extract reason after "PASS - "
                if " - " in output_part:
                    reason = output_part.split(" - ", 1)[1]
                else:
                    reason = "Passed"
            else:
                status = "FAIL"
                # Extract reason after "FAIL - "
                if " - " in output_part:
                    reason = output_part.split(" - ", 1)[1]
                else:
                    reason = "Failed"
            
            return status, reason
        except Exception as e:
            logger.warning(f"Error parsing test output '{test_output}': {e}")
            return "UNKNOWN", "Parse error"
    
    def convert_to_dataframe(self, results: List[Dict[str, Any]]) -> pd.DataFrame:
        """
        Convert evaluation results to pandas DataFrame.
        
        Args:
            results: List of evaluation result dictionaries
            
        Returns:
            DataFrame with columns: mbpp_id, test_case_statement, test_output, reason
        """
        rows = []
        
        for result in results:
            mbpp_id = result.get('mbpp_id', 'unknown')
            test_cases = result.get('test_cases', [])
            test_outputs = result.get('test_outputs', [])
            
            # Ensure we have matching test cases and outputs
            for i, (test_case, test_output) in enumerate(zip(test_cases, test_outputs)):
                status, reason = self.parse_test_output(test_output)
                
                row = {
                    'mbpp_id': mbpp_id,
                    'test_case_statement': test_case,
                    'test_output': status,
                    'reason': reason
                }
                rows.append(row)
            
            # Handle case where there are more test outputs than test cases
            if len(test_outputs) > len(test_cases):
                for i in range(len(test_cases), len(test_outputs)):
                    status, reason = self.parse_test_output(test_outputs[i])
                    row = {
                        'mbpp_id': mbpp_id,
                        'test_case_statement': f"Test case {i+1} (no statement)",
                        'test_output': status,
                        'reason': reason
                    }
                    rows.append(row)
        
        df = pd.DataFrame(rows)
        logger.info(f"Created DataFrame with {len(df)} rows")
        return df
    
    def save_to_excel(self, df: pd.DataFrame, output_path: str = None):
        """
        Save DataFrame to Excel file.
        
        Args:
            df: DataFrame to save
            output_path: Output file path. If None, generates based on input file.
        """
        if output_path is None:
            # Generate output path based on input file
            base_name = os.path.splitext(os.path.basename(self.file_path))[0]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"{base_name}_converted_{timestamp}.xlsx"
        
        try:
            # Create Excel writer with formatting
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Evaluation_Results', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Evaluation_Results']
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add some basic formatting
                from openpyxl.styles import Font, PatternFill
                
                # Header formatting
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Color coding for PASS/FAIL
                pass_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
                fail_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
                
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                    test_output_cell = row[2]  # test_output column (0-indexed)
                    if test_output_cell.value == "PASS":
                        test_output_cell.fill = pass_fill
                    elif test_output_cell.value == "FAIL":
                        test_output_cell.fill = fail_fill
            
            logger.info(f"Excel file saved to: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error saving Excel file: {e}")
            raise
    
    def convert(self, output_path: str = None) -> str:
        """
        Main conversion method.
        
        Args:
            output_path: Output Excel file path (optional)
            
        Returns:
            Path to the created Excel file
        """
        logger.info("Starting conversion process...")
        
        # Load evaluation results
        results = self.load_evaluation_results()
        
        # Convert to DataFrame
        df = self.convert_to_dataframe(results)
        
        # Save to Excel
        output_file = self.save_to_excel(df, output_path)
        
        # Print summary
        logger.info("\n" + "="*50)
        logger.info("CONVERSION SUMMARY")
        logger.info("="*50)
        logger.info(f"Input file: {self.file_path}")
        logger.info(f"Output file: {output_file}")
        logger.info(f"Total rows: {len(df)}")
        logger.info(f"Unique MBPP IDs: {df['mbpp_id'].nunique()}")
        logger.info(f"PASS count: {len(df[df['test_output'] == 'PASS'])}")
        logger.info(f"FAIL count: {len(df[df['test_output'] == 'FAIL'])}")
        
        return output_file

def main():
    """
    Main function to run the conversion.
    """
    parser = argparse.ArgumentParser(description='Convert evaluation results JSON to Excel format')
    parser.add_argument('file_path', type=str,
                       help='Path to the evaluation results JSON file')
    parser.add_argument('--output', type=str, default=None,
                       help='Output Excel file path (optional)')
    parser.add_argument('--debug', action='store_true',
                       help='Enable debug logging')
    
    args = parser.parse_args()
    
    # Set debug logging if requested
    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)
        logger.setLevel(logging.DEBUG)
    
    try:
        # Check if input file exists
        if not os.path.exists(args.file_path):
            logger.error(f"Input file not found: {args.file_path}")
            return
        
        # Initialize converter
        converter = EvalConverter(args.file_path)
        
        # Perform conversion
        output_file = converter.convert(args.output)
        
        logger.info(f"\nConversion completed successfully!")
        logger.info(f"Excel file: {output_file}")
        
    except Exception as e:
        logger.error(f"Error in conversion: {e}")
        raise

if __name__ == "__main__":
    main() 