import json
import os
import pandas as pd
from typing import List, Dict, Any
import argparse
from datetime import datetime

def load_evaluation_results(model_runs_dir: str) -> List[Dict[str, Any]]:
    """
    Load evaluation results from all Model Runs folders.
    
    Args:
        model_runs_dir: Path to the Model Runs directory
        
    Returns:
        List of aggregated evaluation results
    """
    aggregated_results = []
    
    # Get all subdirectories in Model Runs
    model_run_folders = [d for d in os.listdir(model_runs_dir) 
                        if os.path.isdir(os.path.join(model_runs_dir, d)) and d.isdigit()]
    
    # Sort folders numerically
    model_run_folders.sort(key=int)
    
    print(f"Found {len(model_run_folders)} model run folders: {model_run_folders}")
    
    for folder in model_run_folders:
        model_run_id = folder
        evaluations_path = os.path.join(model_runs_dir, folder, "evaluations", "evaluation_results_all.json")
        
        if os.path.exists(evaluations_path):
            print(f"Loading evaluations from {folder}...")
            try:
                with open(evaluations_path, 'r', encoding='utf-8') as f:
                    evaluations = json.load(f)
                
                # Add model_run_id to each evaluation
                for evaluation in evaluations:
                    aggregated_results.append({
                        'model_run_id': model_run_id,
                        'mbpp_id': evaluation.get('mbpp_id'),
                        'pass': evaluation.get('passed', False)
                    })
                
                print(f"  Loaded {len(evaluations)} evaluations from {folder}")
                
            except Exception as e:
                print(f"  Error loading {folder}: {e}")
        else:
            print(f"  No evaluation file found in {folder}")
    
    return aggregated_results

def create_summary_table(results: List[Dict[str, Any]]) -> pd.DataFrame:
    """
    Create a summary table from aggregated results.
    
    Args:
        results: List of evaluation results
        
    Returns:
        DataFrame with columns: model_run_id, mbpp_id, pass
    """
    df = pd.DataFrame(results)
    
    # Ensure proper data types
    df['model_run_id'] = df['model_run_id'].astype(str)
    df['mbpp_id'] = df['mbpp_id'].astype(int)
    df['pass'] = df['pass'].astype(bool)
    
    # Sort by model_run_id and mbpp_id
    df = df.sort_values(['model_run_id', 'mbpp_id'])
    
    return df

def save_results(df: pd.DataFrame, output_dir: str = None):
    """
    Save the aggregated results to JSON and CSV files.
    
    Args:
        df: DataFrame with aggregated results
        output_dir: Output directory (default: current directory)
    """
    if output_dir is None:
        output_dir = "."
    
    os.makedirs(output_dir, exist_ok=True)
    
    # Save as JSON
    json_path = os.path.join(output_dir, "aggregated_evaluations.json")
    results_dict = df.to_dict('records')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(results_dict, f, indent=2, ensure_ascii=False)
    
    # Save as CSV
    csv_path = os.path.join(output_dir, "aggregated_evaluations.csv")
    df.to_csv(csv_path, index=False)
    
    # Save as Excel
    excel_path = os.path.join(output_dir, "aggregated_evaluations.xlsx")
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Evaluations', index=False)
        
        # Get the workbook and worksheet for formatting
        workbook = writer.book
        worksheet = writer.sheets['Evaluations']
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add formatting
        from openpyxl.styles import Font, PatternFill
        
        # Header formatting
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Color coding for pass/fail
        pass_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
        fail_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
        
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            pass_cell = row[2]  # pass column (0-indexed)
            if pass_cell.value == True:
                pass_cell.fill = pass_fill
            elif pass_cell.value == False:
                pass_cell.fill = fail_fill
    
    print(f"Results saved to:")
    print(f"  JSON: {json_path}")
    print(f"  CSV: {csv_path}")
    print(f"  Excel: {excel_path}")
    
    return json_path, csv_path, excel_path

def generate_statistics(df: pd.DataFrame):
    """
    Generate and print statistics from the aggregated results.
    
    Args:
        df: DataFrame with aggregated results
    """
    print("\n" + "="*60)
    print("AGGREGATED EVALUATION STATISTICS")
    print("="*60)
    
    total_evaluations = len(df)
    total_passed = df['pass'].sum()
    total_failed = total_evaluations - total_passed
    overall_pass_rate = total_passed / total_evaluations if total_evaluations > 0 else 0
    
    print(f"Total evaluations: {total_evaluations}")
    print(f"Total passed: {total_passed}")
    print(f"Total failed: {total_failed}")
    print(f"Overall pass rate: {overall_pass_rate:.2%}")
    
    # Statistics by model run
    print(f"\nPass rates by model run:")
    model_stats = df.groupby('model_run_id')['pass'].agg(['count', 'sum']).reset_index()
    model_stats['pass_rate'] = model_stats['sum'] / model_stats['count']
    model_stats = model_stats.sort_values('model_run_id')
    
    for _, row in model_stats.iterrows():
        print(f"  Model Run {row['model_run_id']}: {row['sum']}/{row['count']} ({row['pass_rate']:.2%})")
    
    # Statistics by MBPP ID
    print(f"\nPass rates by MBPP ID (top 10 most common):")
    mbpp_stats = df.groupby('mbpp_id')['pass'].agg(['count', 'sum']).reset_index()
    mbpp_stats['pass_rate'] = mbpp_stats['sum'] / mbpp_stats['count']
    mbpp_stats = mbpp_stats.sort_values('count', ascending=False).head(10)
    
    for _, row in mbpp_stats.iterrows():
        print(f"  MBPP ID {row['mbpp_id']}: {row['sum']}/{row['count']} ({row['pass_rate']:.2%})")

def main():
    """
    Main function to aggregate evaluation results from all Model Runs.
    """
    parser = argparse.ArgumentParser(description='Aggregate evaluation results from all Model Runs folders')
    parser.add_argument('--model-runs-dir', type=str, default='Model Runs',
                       help='Path to Model Runs directory (default: "Model Runs")')
    parser.add_argument('--output-dir', type=str, default=None,
                       help='Output directory for results (default: current directory)')
    
    args = parser.parse_args()
    
    try:
        # Check if Model Runs directory exists
        if not os.path.exists(args.model_runs_dir):
            print(f"Error: Model Runs directory '{args.model_runs_dir}' not found")
            return
        
        print(f"Loading evaluation results from: {args.model_runs_dir}")
        
        # Load all evaluation results
        results = load_evaluation_results(args.model_runs_dir)
        
        if not results:
            print("No evaluation results found!")
            return
        
        print(f"\nLoaded {len(results)} total evaluations")
        
        # Create summary table
        df = create_summary_table(results)
        
        # Save results
        json_path, csv_path, excel_path = save_results(df, args.output_dir)
        
        # Generate statistics
        generate_statistics(df)
        
        # Show sample of the data
        print(f"\nSample of aggregated data:")
        print(df.head(10).to_string(index=False))
        
    except Exception as e:
        print(f"Error: {e}")
        raise

if __name__ == "__main__":
    main() 