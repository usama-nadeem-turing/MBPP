import json
import os
import pandas as pd
from typing import List, Dict, Any

def load_evaluation_results(model_runs_dir: str) -> List[Dict[str, Any]]:
    """Load evaluation results from all Model Runs folders."""
    aggregated_results = []
    
    # Get all subdirectories in Model Runs
    model_run_folders = [d for d in os.listdir(model_runs_dir) 
                        if os.path.isdir(os.path.join(model_runs_dir, d))]
    
    # Sort folders alphabetically
    model_run_folders.sort()
    
    print(f"Found {len(model_run_folders)} model run folders: {model_run_folders}")
    
    for folder in model_run_folders:
        model_run_id = folder
        evaluations_path = os.path.join(model_runs_dir, folder, "evaluations", "evaluation_results_all.json")
        
        if os.path.exists(evaluations_path):
            print(f"Loading evaluations from {folder}...")
            try:
                with open(evaluations_path, 'r', encoding='utf-8') as f:
                    evaluations = json.load(f)
                
                # Add model_run_id to each evaluation
                for evaluation in evaluations:
                    aggregated_results.append({
                        'model_run_id': model_run_id,
                        'mbpp_id': evaluation.get('mbpp_id'),
                        'pass': evaluation.get('passed', False)
                    })
                
                print(f"  Loaded {len(evaluations)} evaluations from {folder}")
                
            except Exception as e:
                print(f"  Error loading {folder}: {e}")
        else:
            print(f"  No evaluation file found in {folder}")
    
    return aggregated_results

def create_pivot_table(results: List[Dict[str, Any]]) -> pd.DataFrame:
    """Create a pivot table with MBPP IDs as rows and model runs as columns."""
    df = pd.DataFrame(results)
    
    # Ensure proper data types
    df['model_run_id'] = df['model_run_id'].astype(str)
    df['mbpp_id'] = df['mbpp_id'].astype(int)
    df['pass'] = df['pass'].astype(bool)
    
    # Create pivot table
    pivot_df = df.pivot(index='mbpp_id', columns='model_run_id', values='pass')
    
    # Rename columns to run_1_pass, run_2_pass, etc.
    pivot_df.columns = [f'run_{col}_pass' for col in pivot_df.columns]
    
    # Reset index to make mbpp_id a column
    pivot_df = pivot_df.reset_index()
    
    # Sort by mbpp_id
    pivot_df = pivot_df.sort_values('mbpp_id')
    
    return pivot_df

def save_pivot_results(df: pd.DataFrame, output_dir: str = "."):
    """Save the pivot table results to JSON, CSV and Excel files."""
    os.makedirs(output_dir, exist_ok=True)
    
    # Save as JSON
    json_path = os.path.join(output_dir, "pivot_evaluations.json")
    results_dict = df.to_dict('records')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(results_dict, f, indent=2, ensure_ascii=False)
    
    # Save as CSV
    csv_path = os.path.join(output_dir, "pivot_evaluations.csv")
    df.to_csv(csv_path, index=False)
    
    # Save as Excel
    excel_path = os.path.join(output_dir, "pivot_evaluations.xlsx")
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Pivot_Evaluations', index=False)
        
        # Get the workbook and worksheet for formatting
        workbook = writer.book
        worksheet = writer.sheets['Pivot_Evaluations']
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 15)  # Smaller width for boolean columns
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add formatting
        from openpyxl.styles import Font, PatternFill
        
        # Header formatting
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Color coding for pass/fail in each run column
        pass_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
        fail_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
        
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            for col_idx, cell in enumerate(row[1:], start=1):  # Skip mbpp_id column
                if cell.value == True:
                    cell.fill = pass_fill
                elif cell.value == False:
                    cell.fill = fail_fill
    
    print(f"Pivot table saved to:")
    print(f"  JSON: {json_path}")
    print(f"  CSV: {csv_path}")
    print(f"  Excel: {excel_path}")
    
    return json_path, csv_path, excel_path

def generate_pivot_statistics(df: pd.DataFrame):
    """Generate and print statistics from the pivot table."""
    print("\n" + "="*60)
    print("PIVOT TABLE STATISTICS")
    print("="*60)
    
    # Get run columns (exclude mbpp_id)
    run_columns = [col for col in df.columns if col.startswith('run_')]
    
    print(f"Total MBPP IDs: {len(df)}")
    print(f"Total model runs: {len(run_columns)}")
    
    # Statistics for each run
    print(f"\nPass rates by model run:")
    for col in run_columns:
        run_num = col.replace('run_', '').replace('_pass', '')
        pass_count = df[col].sum()
        total_count = len(df)
        pass_rate = pass_count / total_count if total_count > 0 else 0
        print(f"  Run {run_num}: {pass_count}/{total_count} ({pass_rate:.2%})")
    
    # Statistics for each MBPP ID
    print(f"\nMBPP IDs with highest pass rates (top 10):")
    df['overall_pass_rate'] = df[run_columns].mean(axis=1)
    top_passing = df.nlargest(10, 'overall_pass_rate')[['mbpp_id', 'overall_pass_rate']]
    
    for _, row in top_passing.iterrows():
        print(f"  MBPP ID {int(row['mbpp_id'])}: {row['overall_pass_rate']:.2%}")
    
    print(f"\nMBPP IDs with lowest pass rates (top 10):")
    bottom_passing = df.nsmallest(10, 'overall_pass_rate')[['mbpp_id', 'overall_pass_rate']]
    
    for _, row in bottom_passing.iterrows():
        print(f"  MBPP ID {int(row['mbpp_id'])}: {row['overall_pass_rate']:.2%}")

def main():
    """Main function to create pivot table from all Model Runs."""
    #model_runs_dir = 'Model Runs'
    model_runs_dir = 'FineTuning\ModelRuns\V2\DoRA_Finetuned'
    
    try:
        # Check if Model Runs directory exists
        if not os.path.exists(model_runs_dir):
            print(f"Error: Model Runs directory '{model_runs_dir}' not found")
            return
        
        print(f"Loading evaluation results from: {model_runs_dir}")
        
        # Load all evaluation results
        results = load_evaluation_results(model_runs_dir)
        
        if not results:
            print("No evaluation results found!")
            return
        
        print(f"\nLoaded {len(results)} total evaluations")
        
        # Create pivot table
        df = create_pivot_table(results)
        
        # Save results
        json_path, csv_path, excel_path = save_pivot_results(df)
        
        # Generate statistics
        generate_pivot_statistics(df)
        
        # Show sample of the data
        print(f"\nSample of pivot table:")
        print(df.head(10).to_string(index=False))
        
        # Show column names
        print(f"\nColumns in pivot table:")
        print(df.columns.tolist())
        
    except Exception as e:
        print(f"Error: {e}")
        raise

if __name__ == "__main__":
    main() 